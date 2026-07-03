#!/usr/bin/env python3
"""Generate PHI-Weaver-Curation-Scorecard.xlsx. Requires openpyxl (`pip install --user openpyxl`); not part of the stdlib-only engine. Edit the `items` list to change the scored rows, then rerun."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

import os
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "PHI-Weaver-Curation-Scorecard.xlsx")

# ---- palette (a chosen slate/teal, not default blue) ----
INK = "1F2A33"       # near-black slate
HEAD = "2E5B62"      # deep teal header
BAND = "EAF0F0"      # pale teal band
META = "F4F1EA"      # warm neutral for metadata
GREEN = "C9E5CB"; AMBER = "F5E4B8"; RED = "F3CBC6"; GREY = "E2E2E2"
WHITE = "FFFFFF"

thin = Side(style="thin", color="B9C4C4")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def st(cell, *, bold=False, size=11, color=INK, fill=None, wrap=False,
       align="left", valign="center", bd=False):
    cell.font = Font(bold=bold, size=size, color=color, name="Calibri")
    cell.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    if bd:
        cell.border = border

wb = openpyxl.Workbook()

# ============================================================ GUIDE
g = wb.active
g.title = "Guide"
g.sheet_view.showGridLines = False
g.column_dimensions["A"].width = 22
g.column_dimensions["B"].width = 96
g.merge_cells("A1:B1")
st(g["A1"], bold=True, size=16, color=WHITE, fill=HEAD)
g["A1"] = "PHI-Canto Curation Scorecard — Guide"
g.row_dimensions[1].height = 26

rows = [
    ("Purpose", "Benchmark the quality of a curation, item by item. Designed so phiweaver "
                "pre-fills the machine-checkable parts (ID validity, ontology-term existence) "
                "and a human curator reviews the judgement calls and completeness."),
    ("Who scores what", "phiweaver auto-check column: filled automatically — does the "
                "identifier/term actually exist and is it current (UniProtKB via query_uniprot; "
                "GO/PHIPO via validate_ontology_ids). Reviewer rating column: the human decides "
                "whether the choice is correct/appropriate."),
    ("Reviewer rating", "Correct — matches the gold standard, ready to submit.\n"
                "Needs improvement — right idea, minor fix (e.g. a less-specific term than ideal).\n"
                "Incorrect — wrong ID, wrong term, or wrong classification.\n"
                "Not applicable — this item is not present in the paper."),
    ("Scoring", "Correct = 1 point, Needs improvement = 0.5, Incorrect = 0. "
                "Not applicable is excluded. Overall accuracy = total points / applicable items "
                "(computed automatically on the Scorecard)."),
    ("Completeness", "Correctness alone is not enough: also record how many curatable items the "
                "paper contained vs how many were captured. A draft that is accurate but misses "
                "half the annotations is not a good draft. Completeness % is computed on the "
                "Scorecard."),
    ("How to use", "1. Copy the 'Scorecard' sheet for each paper (right-click tab > Move or Copy "
                "> tick 'Create a copy'). 2. Fill the header (PMID, date, reviewer). 3. phiweaver "
                "fills the auto-check column; you fill each Reviewer rating from the dropdown and "
                "add a comment for anything not 'Correct'. 4. Copy the paper's overall score + "
                "completeness into the 'Summary' sheet to track over time."),
    ("Note", "A curation scored all-Correct with full completeness is, by definition, a validated "
                "gold-standard example — add it to 07-Standards/curation-examples/."),
]
r = 3
for label, text in rows:
    st(g.cell(r, 1, label), bold=True, color=HEAD, valign="top")
    c = g.cell(r, 2, text); st(c, wrap=True, valign="top")
    g.row_dimensions[r].height = 15 * (1 + text.count("\n") + len(text)//95)
    r += 2

# ============================================================ SCORECARD
s = wb.create_sheet("Scorecard")
s.sheet_view.showGridLines = False
widths = [18, 46, 40, 20, 9, 44]
for i, w in enumerate(widths, 1):
    s.column_dimensions[get_column_letter(i)].width = w

s.merge_cells("A1:F1")
st(s["A1"], bold=True, size=16, color=WHITE, fill=HEAD)
s["A1"] = "PHI-Canto Curation Scorecard"
s.row_dimensions[1].height = 26
s.merge_cells("A2:F2")
st(s["A2"], size=10, color=INK, fill=BAND)
s["A2"] = "phiweaver draft, human-reviewed. See the Guide sheet for the rubric and scoring."

# metadata block
meta = ["Date", "PMID", "Paper title", "Pathogen / host", "Draft produced by", "Reviewer (2nd curator)"]
mr = 4
for i, label in enumerate(meta):
    row = mr + i
    st(s.cell(row, 1, label), bold=True, fill=META, bd=True)
    s.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    st(s.cell(row, 2, ""), fill=WHITE, bd=True)
    for col in range(3, 7):
        s.cell(row, col).border = border

# table header
hr = mr + len(meta) + 1           # header row
headers = ["Annotation level", "Item", "phiweaver auto-check",
           "Reviewer rating", "Points", "Comment"]
for i, h in enumerate(headers, 1):
    st(s.cell(hr, i, h), bold=True, color=WHITE, fill=HEAD, wrap=True, bd=True, align="center")
s.row_dimensions[hr].height = 30

items = [
    ("Entity", "UniProtKB ID correctness", "auto: format valid + accession exists"),
    ("Entity", "Species / strain / cultivar correctness", "— human judgement"),
    ("Gene", "GO / gene annotation correctness (incl. effector GO:0140418)", "auto: GO term exists + current"),
    ("Genotype", "Genotype correctness (allele type + expression level)", "— human judgement"),
    ("Genotype", "Metagenotype & control correctness", "— human judgement"),
    ("Phenotype", "Pathogen phenotype correctness", "auto: PHIPO term exists + current"),
    ("Phenotype", "Host phenotype correctness", "auto: PHIPO term exists + current"),
    ("Phenotype", "Pathogen–host interaction phenotype correctness", "auto: PHIPO term exists + current"),
    ("Detail", "Evidence code correctness", "— human judgement"),
    ("Detail", "Experimental conditions / extensions correctness", "— human judgement"),
    ("Other", "Disease name correctness", "— human judgement"),
    ("Other", "RNA / expression level correctness", "— human judgement"),
    ("Scope?", "Physical / molecular interaction correctness", "confirm PHI-Canto scope; human"),
]
first = hr + 1
for i, (lvl, item, chk) in enumerate(items):
    row = first + i
    band = BAND if i % 2 else WHITE
    st(s.cell(row, 1, lvl), fill=band, bd=True, valign="center")
    st(s.cell(row, 2, item), fill=band, bd=True, wrap=True)
    st(s.cell(row, 3, chk), fill=band, bd=True, wrap=True, size=10, color="55636B")
    st(s.cell(row, 4, ""), fill=WHITE, bd=True, align="center")     # rating (dropdown)
    pc = s.cell(row, 5)                                             # points (formula)
    pc.value = (f'=IF($D{row}="Correct",1,IF($D{row}="Needs improvement",0.5,'
                f'IF($D{row}="Incorrect",0,"")))')
    st(pc, bd=True, align="center")
    st(s.cell(row, 6, ""), fill=WHITE, bd=True, wrap=True)
    s.row_dimensions[row].height = 26
last = first + len(items) - 1

# dropdown + conditional colours on the rating column
dv = DataValidation(type="list",
                    formula1='"Correct,Needs improvement,Incorrect,Not applicable"',
                    allow_blank=True)
s.add_data_validation(dv)
dv.add(f"D{first}:D{last}")
rng = f"D{first}:D{last}"
for val, col in (("Correct", GREEN), ("Needs improvement", AMBER),
                 ("Incorrect", RED), ("Not applicable", GREY)):
    s.conditional_formatting.add(rng, CellIsRule(
        operator="equal", formula=[f'"{val}"'], fill=PatternFill("solid", fgColor=col)))

# score summary
sr = last + 2
def summary(label, formula, pct=False, bold=False):
    global sr
    st(s.cell(sr, 4, label), bold=True, fill=META, bd=True, align="right")
    c = s.cell(sr, 5, formula); st(c, bold=bold, fill=WHITE, bd=True, align="center")
    if pct:
        c.number_format = "0%"
    sr += 1
summary("Applicable items", f"=COUNT(E{first}:E{last})")
summary("Points", f"=SUM(E{first}:E{last})")
summary("Overall accuracy", f'=IF(E{sr-2}=0,"",E{sr-1}/E{sr-2})', pct=True, bold=True)

# completeness block
cr = sr + 1
st(s.cell(cr, 1, "Completeness"), bold=True, color=WHITE, fill=HEAD, bd=True)
s.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=3)
for col in (2, 3):
    s.cell(cr, col).fill = PatternFill("solid", fgColor=HEAD)
    s.cell(cr, col).border = border
comp = [
    ("Curatable items in the paper", 0),
    ("Items captured in the draft", 0),
    ("Items missed", None),          # formula
    ("Completeness", None),          # formula %
]
cr += 1
base = cr
for i, (label, default) in enumerate(comp):
    row = cr + i
    st(s.cell(row, 1, label), bold=True, fill=META, bd=True)
    s.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    s.cell(row, 2).border = border; s.cell(row, 2).fill = PatternFill("solid", fgColor=META)
    cell = s.cell(row, 3)
    if label == "Items missed":
        cell.value = f"=MAX(0,C{base}-C{base+1})"
    elif label == "Completeness":
        cell.value = f'=IF(C{base}=0,"",C{base+1}/C{base})'
        cell.number_format = "0%"
    else:
        cell.value = default
    st(cell, bd=True, align="center", bold=(label == "Completeness"))

s.freeze_panes = f"A{hr+1}"

# ============================================================ SUMMARY
sm = wb.create_sheet("Summary")
sm.sheet_view.showGridLines = False
for i, w in enumerate([16, 12, 22, 16, 14, 50], 1):
    sm.column_dimensions[get_column_letter(i)].width = w
sm.merge_cells("A1:F1")
st(sm["A1"], bold=True, size=14, color=WHITE, fill=HEAD)
sm["A1"] = "Benchmark Summary — one row per scored paper"
sm.row_dimensions[1].height = 24
hdr = ["PMID", "Date", "Reviewer", "Overall accuracy", "Completeness", "Notes"]
for i, h in enumerate(hdr, 1):
    st(sm.cell(3, i, h), bold=True, color=WHITE, fill=HEAD, bd=True, align="center")
for r_ in range(4, 30):
    for c_ in range(1, 7):
        cell = sm.cell(r_, c_); cell.border = border
        if c_ in (4, 5):
            cell.number_format = "0%"
        if r_ % 2 == 0:
            cell.fill = PatternFill("solid", fgColor=BAND)
sm.freeze_panes = "A4"

import os
os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print("wrote", OUT)
