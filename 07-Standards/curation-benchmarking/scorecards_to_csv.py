#!/usr/bin/env python3
"""
scorecards_to_csv.py — export filled benchmarking scorecards to the scores CSV that
phiweaver.benchmark_report consumes.

After a curator has filled a scorecard's **Reviewer-rating** column and the **completeness**
block, this reads each `.xlsx` and writes one CSV row per paper:
    paper, group, model, curatable, captured, <item>, <item>, ...
Curated papers are passed positionally; held-out control papers with --control. Ratings are read
verbatim ("Not applicable" -> "N/A"); accuracy is computed downstream by benchmark_report.

Requires openpyxl. Label-based cell lookup, so it survives layout tweaks to the .xlsx.

Usage:
    python3 scorecards_to_csv.py curated/*.xlsx --control control/*.xlsx --out scores.csv
    python3 -m phiweaver.benchmark_report scores.csv --out benchmark-report.html
"""
from __future__ import annotations

import argparse
import csv
from pathlib import Path

import openpyxl

ITEM_HEADER = "annotation level"          # col-A marker for the item-table header row
PAPER_LABELS = ("paper title", "pmid")    # col-A labels to source the paper name from
MODEL_LABEL = "model"                     # col-A label for the drafting-model provenance row
COMPLETENESS = {"curatable items in the paper": "curatable",
                "items captured in the draft": "captured"}
RATING_NORM = {"not applicable": "N/A", "n/a": "N/A"}


def _norm(v) -> str:
    v = str(v or "").strip()
    return RATING_NORM.get(v.lower(), v)


def read_scorecard(path, group: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    s = wb["Scorecard"]
    a_row, header = {}, None
    for r in range(1, s.max_row + 1):
        a = str(s.cell(r, 1).value or "").strip()
        if a:
            a_row.setdefault(a.lower(), r)
            if a.lower() == ITEM_HEADER and header is None:
                header = r

    paper = ""
    for lbl in PAPER_LABELS:
        if lbl in a_row:
            v = str(s.cell(a_row[lbl], 2).value or "").strip()
            if v and v.lower() != "to confirm":
                paper = v
                break
    if not paper:
        paper = Path(path).stem

    model = ""
    if MODEL_LABEL in a_row:
        model = str(s.cell(a_row[MODEL_LABEL], 2).value or "").strip()

    # item rows = the contiguous rows after the header that have a non-empty Item (col B)
    ratings, started = {}, False
    if header:
        for r in range(header + 1, s.max_row + 1):
            item = str(s.cell(r, 2).value or "").strip()
            if item:
                ratings[item] = _norm(s.cell(r, 4).value)   # col D = Reviewer rating
                started = True
            elif started:
                break

    comp = {"curatable": "", "captured": ""}
    for lbl, key in COMPLETENESS.items():
        if lbl in a_row:
            v = s.cell(a_row[lbl], 3).value                  # col C = value
            comp[key] = "" if v is None else v
    return {"paper": paper, "group": group, "model": model, "ratings": ratings, **comp}


def to_csv(records, out):
    items = []
    for rec in records:
        for it in rec["ratings"]:
            if it not in items:
                items.append(it)
    with open(out, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["paper", "group", "model", "curatable", "captured"] + items)
        for rec in records:
            w.writerow([rec["paper"], rec["group"], rec.get("model", ""),
                        rec["curatable"], rec["captured"]]
                       + [rec["ratings"].get(it, "") for it in items])


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(
        description="Export filled scorecards to the benchmark_report scores CSV.")
    ap.add_argument("scorecards", nargs="*", help="filled scorecard .xlsx (curated group)")
    ap.add_argument("--control", nargs="*", default=[], help="held-out control scorecards")
    ap.add_argument("--out", default="scores.csv")
    args = ap.parse_args(argv)
    if not args.scorecards and not args.control:
        ap.error("provide at least one scorecard")
    records = ([read_scorecard(p, "curated") for p in args.scorecards]
               + [read_scorecard(p, "control") for p in args.control])
    to_csv(records, args.out)
    n_items = len({it for r in records for it in r["ratings"]})
    print(f"wrote {args.out} ({len(records)} paper(s), {n_items} item(s))")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
