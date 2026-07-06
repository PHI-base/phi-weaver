#!/usr/bin/env python3
"""
fill_scorecard.py — pre-fill the benchmarking scorecard from a phiweaver draft.

A phiweaver curation draft carries a machine-readable auto-check block (a fenced ```json
block; see 07-Standards/curation-examples/_TEMPLATE.md). This reads that block and writes a
copy of PHI-Weaver-Curation-Scorecard.xlsx with the **header** and the **phiweaver auto-check
column** pre-filled — leaving the human Reviewer-rating column blank by design.

Single or batch:
    python3 fill_scorecard.py draft.md
    python3 fill_scorecard.py draft.md --out /path/scorecard.xlsx
    python3 fill_scorecard.py active/*-phiweaver-DRAFT.md      # one scorecard per draft

Requires openpyxl (`pip install --user openpyxl`); not part of the stdlib-only engine.
Row/label lookups are by text, so it survives layout tweaks to the .xlsx.
"""
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

import openpyxl

TEMPLATE_DEFAULT = Path(__file__).with_name("PHI-Weaver-Curation-Scorecard.xlsx")

# draft meta key -> scorecard metadata row label (matched in column A)
META_LABELS = {
    "date": "Date", "pmid": "PMID", "paper": "Paper title",
    "system": "Pathogen / host", "draft_by": "Draft produced by",
    "model": "Model",
}
# auto_check key -> a substring of the scorecard Item label (matched in column B)
KEY_TO_LABEL = {
    "uniprot_id": "uniprotkb id",
    "species_strain_cultivar": "species / strain",
    "go_gene_annotation": "go / gene annotation",
    "genotype": "genotype correctness",
    "metagenotype_control": "metagenotype & control",
    "pathogen_phenotype": "pathogen phenotype",
    "host_phenotype": "host phenotype",
    "interaction_phenotype": "interaction phenotype",
    "evidence_code": "evidence code",
    "conditions_extensions": "conditions / extensions",
    "disease_name": "disease name",
    "rna_expression_level": "rna / expression",
    "physical_interaction": "physical / molecular interaction",
}

_JSON_BLOCK = re.compile(r"```json\s*(.*?)```", re.DOTALL)


def extract_record(md_text: str):
    """Parse the first ```json ... ``` block in the draft, or return None."""
    m = _JSON_BLOCK.search(md_text)
    if not m:
        return None
    return json.loads(m.group(1))


def fill(draft_path, template, out_path):
    rec = extract_record(Path(draft_path).read_text(encoding="utf-8"))
    if rec is None:
        raise SystemExit(f"no ```json auto-check block found in {draft_path}")
    wb = openpyxl.load_workbook(template)        # fill a fresh copy of the template
    s = wb["Scorecard"]

    # index label -> row for column A (metadata labels) and column B (item labels)
    a_rows, b_rows = {}, {}
    for r in range(1, s.max_row + 1):
        a, b = s.cell(r, 1).value, s.cell(r, 2).value
        if isinstance(a, str) and a.strip():
            a_rows.setdefault(a.strip().lower(), r)
        if isinstance(b, str) and b.strip():
            b_rows.setdefault(b.strip().lower(), r)

    for key, label in META_LABELS.items():
        val = rec.get("meta", {}).get(key)
        row = a_rows.get(label.lower())
        if val and row:
            s.cell(row, 2).value = val

    for key, val in rec.get("auto_check", {}).items():
        sub = KEY_TO_LABEL.get(key)
        if not sub or not val:
            continue
        row = next((rr for lbl, rr in b_rows.items() if sub in lbl), None)
        if row:
            s.cell(row, 3).value = val            # column C = phiweaver auto-check

    out_path = Path(out_path)
    try:
        if out_path.exists():
            out_path.unlink()                     # z: mount: remove before rewriting
        wb.save(str(out_path))
    except PermissionError:
        raise SystemExit(
            f"cannot write {out_path} — is it open in Excel? Close it and retry.")
    return out_path


def default_out(draft_path):
    p = Path(draft_path)
    stem = p.stem.replace("-phiweaver-DRAFT", "").replace("-DRAFT", "")
    return p.with_name(f"{stem}-scorecard-PREFILLED.xlsx")


def main(argv=None):
    ap = argparse.ArgumentParser(
        description="Pre-fill benchmarking scorecard(s) from phiweaver draft(s).")
    ap.add_argument("drafts", nargs="+",
                    help="draft .md file(s) containing a ```json auto-check block")
    ap.add_argument("--template", default=str(TEMPLATE_DEFAULT))
    ap.add_argument("--out", help="output path (single draft only)")
    args = ap.parse_args(argv)
    if args.out and len(args.drafts) != 1:
        ap.error("--out is only valid with a single draft")
    for d in args.drafts:
        out = args.out or default_out(d)
        fill(d, args.template, out)
        print("wrote", out)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
